import openpyxl

path="C:\\Users\\User\\Desktop\\DDTest.xlsx"

workBook=openpyxl.load_workbook(path) #creating a workbook object

sheet=workBook.active #getting the active sheet, mostly when there is only one sheet active in the workbook

#an alternative method is to get the sheet by its name when there are more than one sheet, as follows
#sheet=workBook.get_sheet_by_name("sheet1")

rows=sheet.max_row #to get the number of rows off the worksheet
cols=sheet.max_column #to get the number of columns off the worksheet

print(rows)  #14
print(cols)  #4

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="        ")

    print()

